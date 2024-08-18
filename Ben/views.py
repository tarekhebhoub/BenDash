# views.py
from rest_framework import viewsets # type: ignore
from .models import Customer, Product, Invoice, Payment,InvoiceItem
from .serializers import CustomerSerializer, ProductSerializer, InvoiceSerializer, PaymentSerializer,InvoiceItemSerializer
from rest_framework.decorators import api_view, permission_classes# type: ignore
from rest_framework.permissions import IsAuthenticated # type: ignore
from rest_framework.response import Response# type: ignore
from rest_framework.authtoken.models import Token# type: ignore
from rest_framework import status# type: ignore
from datetime import datetime
from dateutil.relativedelta import relativedelta

from django.db.models import Case, When, Value, IntegerField
from django.db.models.functions import ExtractDay
from django.utils import timezone

# Get the current day of the month
import csv
from django.http import HttpResponse


from django.http import JsonResponse
from django.db.models import Sum



from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter


# Input date string
def changeDate(date_str):
    if date_str == None:
        return None  

# Parse the input string to a datetime object
# Note: The 'Z' at the end of the string indicates UTC time
    date_obj = datetime.strptime(date_str, "%Y-%m-%dT%H:%M:%SZ")

# Format the datetime object to the desired string format
    formatted_date = date_obj.strftime("%Y-%m-%d")
    return formatted_date

class CustomerViewSet(viewsets.ModelViewSet):
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer

class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer

class InvoiceViewSet(viewsets.ModelViewSet):
    queryset = Invoice.objects.all()
    serializer_class = InvoiceSerializer

class PaymentViewSet(viewsets.ModelViewSet):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer


def increment_date_by_months(date_str, months):
    date_format = '%Y-%m-%dT%H:%M:%S.%fZ'

    # Parse the input date string to a datetime object
    date_obj = datetime.strptime(date_str, "%Y-%m-%d")
    
    # Add the specified number of months to the date object
    new_date = date_obj + relativedelta(months=months)
    
    # Return the new date as a string in the desired format
    return new_date.strftime("%Y-%m-%d")
# Example usage







@api_view(['Get'])  # Use the appropriate HTTP method for your API
@permission_classes([IsAuthenticated])
def GetVentes(request):
    # invoices=Invoice.objects.all()

    # Get the current day of the month
    today_day = timezone.now().day

    # Query invoices with conditional ordering based on the day of the month
    invoices = Invoice.objects.annotate(
        day_of_date_dub=ExtractDay('date_dub')  # Extract the day of the month from date_dub
    ).order_by(
        Case(
            When(day_of_date_dub=today_day, then=Value(0)),  # Invoices with today's day first
            default=Value(1),                                # Others come later
            output_field=IntegerField()                      # Field type for ordering
        ),
        'date_dub'  # Secondary ordering by date_dub if needed
    )

    print(invoices)


    ventes=InvoiceSerializer(invoices,many=True)
    data=[]
    for invoice in ventes.data:
        date_dub=datetime.strptime(invoice['date_dub'], "%Y-%m-%dT%H:%M:%SZ")
        date_dub=date_dub.strftime("%Y-%m-%d")
        date_fin = increment_date_by_months(date_dub, int(invoice['installment_period'])-1)
        invoice["date_fin"]=date_fin
        invoice["date_dub"]=date_dub
        created_at=datetime.strptime(invoice['created_at'], "%Y-%m-%dT%H:%M:%SZ")
        created_at=created_at.strftime("%Y-%m-%d")
        invoice['created_at']=created_at
        customer=Customer.objects.get(id=invoice['customer'])
        invoice['customer_Name']=customer.name
        invoice['customer_id']=customer.id
        invoice['customer_CCP']=customer.ccp
        invoice['phone']=customer.phone

        payments=Payment.objects.filter(invoice=invoice['id'])
        paymentSerializer=PaymentSerializer(payments,many=True)
        
        prix_Vente=0
        invoiceItems=InvoiceItem.objects.filter(invoice=invoice['id'])
        Items_Serializer=InvoiceItemSerializer(invoiceItems,many=True)
        for item in Items_Serializer.data:
            prix_Vente=prix_Vente+float(item['total_price'])
        invoice['prix_Vente']="{:.2f}".format(prix_Vente)
        prix_Vente=prix_Vente-(float(invoice['remise'])+float(invoice['init_amount']))
        invoice['prix_Final']="{:.2f}".format(prix_Vente)
        mententPrelvement=prix_Vente/int(invoice['installment_period'])
        invoice['mententPrelvement']="{:.2f}".format(mententPrelvement)
        total=0
        for payment in paymentSerializer.data:

            amount = payment['amount'] or 0  

            total=total+float(amount)
        rest=prix_Vente-total
        invoice["rest"]="{:.2f}".format(rest)
        data.append(invoice)    
    return Response(data)





@api_view(['Get'])  # Use the appropriate HTTP method for your API
@permission_classes([IsAuthenticated])
def GetVenteDetails(request,pk):
    invoices=Invoice.objects.get(id=pk)
    vente=InvoiceSerializer(invoices)
    invoice=vente.data
    date_dub=datetime.strptime(invoice['date_dub'], "%Y-%m-%dT%H:%M:%SZ")
    date_dub=date_dub.strftime("%Y-%m-%d")
    date_fin = increment_date_by_months(date_dub, int(invoice['installment_period']))
    invoice["date_fin"]=date_fin
    invoice["date_dub"]=date_dub
    created_at=datetime.strptime(invoice['created_at'], "%Y-%m-%dT%H:%M:%SZ")
    created_at=created_at.strftime("%Y-%m-%d")
    invoice['created_at']=created_at
    customer=Customer.objects.get(id=invoice['customer'])
    invoice['customer_Name']=customer.name
    invoice['customer_CCP']=customer.ccp

    payments=Payment.objects.filter(invoice=invoice['id'])
    paymentSerializer=PaymentSerializer(payments,many=True)
    
    
    prix_Vente=0
    invoiceItems=InvoiceItem.objects.filter(invoice=invoice['id'])
    Items_Serializer=InvoiceItemSerializer(invoiceItems,many=True)
    for item in Items_Serializer.data:
        prix_Vente=prix_Vente+float(item['total_price'])
    invoice['prix_Vente']="{:.2f}".format(prix_Vente)
    prix_Vente=prix_Vente-(float(invoice['remise'])+float(invoice['init_amount']))

    invoice['prix_Final']="{:.2f}".format(prix_Vente)
    mententPrelvement=prix_Vente/int(invoice['installment_period'])
    mententPrelvement="{:.2f}".format(mententPrelvement)
    invoice['mententPrelvement']=mententPrelvement
    total=0
    for payment in paymentSerializer.data:
        amount = payment['amount'] or 0  # If amount is None, use 0

        total=total+float(amount)
    rest=float(invoice['prix_Final'])- float(total)
    invoice["rest"]="{:.2f}".format(rest)
    products=InvoiceItem.objects.filter(invoice=invoice['id'])
    products=InvoiceItemSerializer(products,many=True)
    invoice['products']=products.data

    installments=[]
    i=0
    for payment in paymentSerializer.data:
        payment['mententPrelvement']=mententPrelvement
        payment['dateVente']=increment_date_by_months(invoice['date_dub'],i)
        payment['payment_date']=changeDate(payment['payment_date'])
        print(payment)

        i+=1
        installments.append(payment)
    # print(installments)
    invoice['installments']=installments

    return Response(invoice)

''' 
    if len(installments)<int(invoice['installment_period']):
        x=int(invoice['installment_period'])-len(installments)
        for j in range(x):
            payment = {
                'invoice': invoice['id'],
                'mententPrelvement': mententPrelvement,
                'amount': None,
                'payment_date': None,
                'dateVente': increment_date_by_months(date_dub, i + j)
            }
            installments.append(payment)

    
'''
    # print(installments)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def Create_invoice_with_products(request):
    """
    Create a new invoice and add products to it, all from a single request.
    """
    invoice_data = request.data

    # Extract the products data from the request
    products_data = invoice_data.pop('products', [])
    print(invoice_data)


    # # Remove 'total_price' from each product data
    # for product_data in products_data:
    #     if 'total_price' in product_data:
    #         del product_data['total_price']
    # print(products_data)

    # Create the invoice
    invoice_serializer = InvoiceSerializer(data=invoice_data)
    if invoice_serializer.is_valid():
        invoice = invoice_serializer.save()  # Save the invoice and get the instance
        print("-------this done--------------")
        # Process the products
        for product_data in products_data:
            product_id = product_data.get('product_id')
            quantity = product_data.get('quantity', 1)
            total_price=product_data.get('total_price')

            if not product_id:
                return Response({'error': 'Product ID is required'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                product = Product.objects.get(id=product_id)
                product.quantity=product.quantity-int(quantity)
                product.save()
            except Product.DoesNotExist:
                return Response({'error': f'Product with ID {product_id} not found'}, status=status.HTTP_404_NOT_FOUND)

            # Create the InvoiceItem
            InvoiceItem.objects.create(
                invoice=invoice,
                product=product,
                quantity=quantity,
                total_price=total_price
            )
        # Serialize the updated invoice


        nmbMonth=int(invoice_data['installment_period'])
        for i in range (nmbMonth):
            new_item = Payment.objects.create(
                amount=None,
                payment_date=None,
                invoice=invoice,
                # Add other fields as needed
            )



        updated_invoice_serializer = InvoiceSerializer(invoice)
        return Response(updated_invoice_serializer.data, status=status.HTTP_201_CREATED)
    else:
        print(invoice_serializer.errors)
        return Response(invoice_serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# def postPaiment(request):
#     serializer=PaymentSerializer(request.data)
#     if serializer.is_valid():
#         serializer.save()
#         return {"Done"}
#     else:
#         return serializer.error

@api_view(['Get'])  # Use the appropriate HTTP method for your API
@permission_classes([IsAuthenticated])
def tryToken(request):
    return Response({'tarek'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout(request):
    try:
        # Get the user's token
        token = Token.objects.get(user=request.user)
        # Delete the token to effectively log out the user
        token.delete()
        return Response({"detail": "Successfully logged out."}, status=status.HTTP_200_OK)
    except Token.DoesNotExist:
        return Response({"detail": "Token not found."}, status=status.HTTP_400_BAD_REQUEST)










def PrixFinal(id):
    invoices=Invoice.objects.get(id=id)
    invoice=InvoiceSerializer(invoices)
    invoice=invoice.data
    invoiceItems=InvoiceItem.objects.filter(invoice=id)


    prix_Vente=0
    Items_Serializer=InvoiceItemSerializer(invoiceItems,many=True)
    for item in Items_Serializer.data:
        prix_Vente=prix_Vente+float(item['total_price'])
    prix_Vente=prix_Vente-(float(invoice['remise'])+float(invoice['init_amount']))
    return prix_Vente;
    # invoice['prix_Final']="{:.2f}".format(prix_Vente)




@api_view(['Get'])  # Use the appropriate HTTP method for your API
@permission_classes([IsAuthenticated])

def export_clients_with_rest(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="clients_with_restDate.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Clients Report'

  

    # Define styles
    bold_font = Font(bold=True, size=16)
    border_style = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))

    # Add title row
    title = f' قائمة الزبائن  للأقساط الغير منتهية '
    ws.merge_cells('A1:E1')  # Adjust columns based on your needs
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = bold_font
    title_cell.border = border_style

    # Add header row
    headers = ['Name', 'CCP', 'Phone', 'Prix Ventes', 'Rest']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = bold_font
        cell.border = border_style

    # Adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 16  # Adjust width as needed


    invoices=Invoice.objects.all()
        
    row_num=3
    for invoice in invoices:
        payments = Payment.objects.filter(invoice=invoice)
        paid_amount = sum(payment.amount for payment in payments)
        prixFinal = PrixFinal(invoice.id)
        rest = prixFinal - float(paid_amount)

        if rest != 0:
            client = invoice.customer
            
            ws.append([client.name, client.ccp, '0'+str(client.phone), prixFinal, rest])

            # Apply border to all cells in the row
            for cell in ws[row_num]:
                cell.border = border_style
            row_num += 1

    wb.save(response)
    return response


@api_view(['Get'])  # Use the appropriate HTTP method for your API
@permission_classes([IsAuthenticated])
def export_clients_with_rest_with_date(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="clients_with_restDate.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Clients Report'

    dateFrom = int(request.GET.get('dateFrom'))
    dateTo = int(request.GET.get('dateTo'))

    # Define styles
    bold_font = Font(bold=True, size=16)
    border_style = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))

    # Add title row
    title = f'قائمة الزبائن لدفع الأقساط من {dateFrom} إلى {dateTo}'
    ws.merge_cells('A1:E1')  # Adjust columns based on your needs
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = bold_font
    title_cell.border = border_style

    # Add header row
    headers = ['Name', 'CCP', 'Phone', 'Prix Ventes', 'Rest']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = bold_font
        cell.border = border_style

    # Adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 16  # Adjust width as needed

    # Fetch invoices and check for zero rest

    print(dateFrom)
    print(dateTo)

    invoices=Invoice.objects.all()
    filtered_invoices=[]
    for invoice in invoices:
        print(invoice.date_dub.day)
        if invoice.date_dub.day >= dateFrom and invoice.date_dub.day <= dateTo:
            filtered_invoices.append(invoice)
    row_num = 3
    print(filtered_invoices)
    for invoice in filtered_invoices:
        payments = Payment.objects.filter(invoice=invoice)
        paid_amount = sum(payment.amount for payment in payments)
        prixFinal = PrixFinal(invoice.id)
        rest = prixFinal - float(paid_amount)

        if rest != 0:
            client = invoice.customer
            
            ws.append([client.name, client.ccp, '0'+str(client.phone), prixFinal, rest])

            # Apply border to all cells in the row
            for cell in ws[row_num]:
                cell.border = border_style
            row_num += 1

    wb.save(response)
    return response
@api_view(['Get'])  # Use the appropriate HTTP method for your API
@permission_classes([IsAuthenticated])

def dashboardData(request):
    num_clients = Customer.objects.count()
    num_ventes = Invoice.objects.count()
   

    invoices = Invoice.objects.all()
   
    restTotal=0
    for invoice in invoices:
        payments = Payment.objects.filter(invoice=invoice)
        paid_amount = sum(payment.amount for payment in payments)
        prixFinal=PrixFinal(invoice.id)

        rest =  prixFinal- float(paid_amount)
        restTotal=restTotal+rest



    total_mentant_payed = Payment.objects.aggregate(total_paid=Sum('amount'))['total_paid'] or 0

    data = {
        'numClients': num_clients,
        'numVentes': num_ventes,
        'restAPayer': float(restTotal),
        'mentantPayed': float(total_mentant_payed),
    }

    return JsonResponse(data)

